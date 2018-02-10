from openpyxl import load_workbook
from openpyxl import Workbook
import os
import json


# 常量
display_dic = {
    'Order ID': 3,
    'Customer ID': 2,
    'SKU': 4,
    'Product Name': 5,
    'Address': 7,
    'Customer Name': 1,
    'Phone': 10
    }


default_fill_str = ''
default_fill_total_money = '20'
default_fill_total_real_money = '20'
default_fill_real_money = '15'
default_province = '中'
default_city = '国'
default_county = '区'

#用户自定义的格式，读入 JSON
SKU_STR = {}
#快递按照地区分区
EXPRESS_REGION = {}
#快递按照数量分区
EXPRESS_COUNT = {}


now_dir_for_load_file = os.path.dirname(os.path.abspath(__file__))

def load_config():
    print('now', now_dir_for_load_file)
    # 设置以utf-8解码模式读取文件，encoding参数必须设置，否则默认以gbk模式读取文件，当文件中包含中文时，会报错
    with open(os.path.join(now_dir_for_load_file,"config.json"), encoding='utf-8') as f:
        return json.load(f)


def load_express_region():
    # 设置以utf-8解码模式读取文件，encoding参数必须设置，否则默认以gbk模式读取文件，当文件中包含中文时，会报错
    with open(os.path.join(now_dir_for_load_file, "region.json"), encoding='utf-8') as f:
        return json.load(f)


def load_express_count():
    # 设置以utf-8解码模式读取文件，encoding参数必须设置，否则默认以gbk模式读取文件，当文件中包含中文时，会报错
    with open(os.path.join(now_dir_for_load_file, "region-by-count.json"), encoding='utf-8') as f:
        return json.load(f)


def pre_do(sheet):
    # 处理提取字段在原始表中的序号， 初始化 display_dic
    for i, cell in enumerate(sheet['1']):
        if cell.value in display_dic.keys():
            display_dic[cell.value] = i


def get_num(str):
    # 获取数量，从 ："内蒙古【胡汉和亲】牛肉干 正宗内蒙风干制作工艺 风干牛肉干200克/袋 * 3规格:原味
    num = '0'
    try:
        num = str.split('规格')[0].split('*')[-1].strip()
    except:
        pass
    return num


def do_row(row):
    """
    从读入文档的每行中提取需要信息
    :param row: 读入文档的行
    :return: 返回处理提取结果
    """
    phones = row[display_dic['Phone']].value.split('/')

    headlines = ''
    SKUS = row[display_dic['SKU']].value.split('\n')
    # print(SKUS, phones)
    product_names = row[display_dic['Product Name']].value.split('\n')

    for i in range(len(SKUS)):
        if SKUS[i] != '':
            headlines += SKU_STR[SKUS[i]] + "*" + get_num(product_names[i]) +"    "

    get_raw_info = {
        'order_id': row[display_dic['Order ID']].value,
        'buyer_nike': row[display_dic['Customer ID']].value,
        # 'headline': SKU_STR[row[display_dic['SKU']].value.strip('\n')],
        'headline': headlines,
        'product_num': get_num(row[display_dic['Product Name']].value),
        'receiver_name': row[display_dic['Customer Name']].value,
        'fixed_phone': phones[0],
        'mobile_phone': phones[1] if len(phones) > 1 else phones[0],
        'detail_address': row[display_dic['Address']].value,
    }

    return get_raw_info


def do_row_count_sort(row, ll):
    """
    从读入文档的每行中提取需要信息, 根据 ll　判断是否属于这个分类
    :param row: 读入文档的行
    :return: 返回处理提取结果
    """
    phones = row[display_dic['Phone']].value.split('/')

    headlines = ''
    SKUS = row[display_dic['SKU']].value.split('\n')
    # print(SKUS, phones)
    product_names = row[display_dic['Product Name']].value.split('\n')

    def sku_in_list():
        for i in SKUS:
            if i in ll:
                return True
        return False

    for i in range(len(SKUS)):
        if SKUS[i] != '':
            headlines += SKU_STR[SKUS[i]] + "*" + get_num(product_names[i]) +" "

    get_raw_info = {
        'order_id': row[display_dic['Order ID']].value,
        'buyer_nike': row[display_dic['Customer ID']].value,
        # 'headline': SKU_STR[row[display_dic['SKU']].value.strip('\n')],
        'headline': headlines,
        'product_num': get_num(row[display_dic['Product Name']].value),
        'receiver_name': row[display_dic['Customer Name']].value,
        'fixed_phone': phones[0],
        'mobile_phone': phones[1] if len(phones) > 1 else phones[0],
        'detail_address': row[display_dic['Address']].value,
    }

    # if sku_in_list() or int(get_num(row[display_dic['Product Name']].value)) > 1:

    def len_list_contain_emptustr(strl):
        return len([x for x in strl if x != '' ])


    if len_list_contain_emptustr(SKUS) > 1 or sku_in_list() or int(get_num(row[display_dic['Product Name']].value)) > 1:
        return True, get_raw_info
    else:
        return False, get_raw_info



def generate_out_row(get_raw_info):
    """
    生成打印格式内容
    :param get_raw_info: 所需字段 
    :return: 
    """
    r =[get_raw_info['order_id'], get_raw_info['buyer_nike'],
        default_fill_total_money, default_fill_total_real_money,default_fill_str, default_fill_str,
        get_raw_info['headline'],
        default_fill_str, default_fill_str, default_fill_str, default_fill_str,
        get_raw_info['product_num'],
        default_fill_str, default_fill_str, default_fill_real_money,
        get_raw_info['receiver_name'], get_raw_info['fixed_phone'], get_raw_info['mobile_phone'],
        default_province, default_city, default_county,
        get_raw_info['detail_address'],
        default_fill_str, default_fill_str]

    return r


def readxl_split_by_express_region(file_name):
    wb = load_workbook(file_name)
    sheet = wb.get_sheet_by_name("Worksheet")

    pre_do(sheet)

    #这个数据字典按照，key是快递，values是列表构成存放数据
    data_dic ={}
    for k in EXPRESS_REGION.keys():
        data_dic[k] = []

    # 总表不分快递公司
    data_total_table = []
    first = False
    for row in sheet.rows:
        #需要处理一下，因为表格后面有其他多余数据
        try:
            if type(row[0].value) == 'NoneType' or row[0].value.strip('\n').strip() == '':
                continue
        except:
            continue
        if first:
            # 快递分类，打印行中 21 列 是 地区
            dealed_data = generate_out_row(do_row(row))
            data_total_table.append(dealed_data)
            addr = dealed_data[21].split(' ')
            for k in EXPRESS_REGION.keys():
                for c in EXPRESS_REGION[k]:
                    if c in addr:
                        data_dic[k].append(dealed_data)
                        break
        first = True

    wb.close()
    return data_dic, data_total_table


def readxl_split_by_express_by_count(file_name):
    wb = load_workbook(file_name)
    sheet = wb.get_sheet_by_name("Worksheet")

    pre_do(sheet)

    data_dic = {"韵达":[], "邮政":[]}

    # 总表不分快递公司
    data_total_table = []
    first = False
    for row in sheet.rows:
        #需要处理一下，因为表格后面有其他多余数据
        try:
            if type(row[0].value) == 'NoneType' or row[0].value.strip('\n').strip() == '':
                continue
        except:
            continue
        if first:
            # 按照数量啥的进行分类，特殊处理得函数
            yundaed, done_row = do_row_count_sort(row, EXPRESS_COUNT['韵达'])
            dealed_data = generate_out_row(done_row)
            data_total_table.append(dealed_data)
            if yundaed:
                data_dic['韵达'].append(dealed_data)
            else:
                data_dic['邮政'].append(dealed_data)
        first = True

    wb.close()
    return data_dic, data_total_table


def writexl(data_dic, base_dir):
    out_path = []
    header = ["订单号(必填)","买家昵称(必填)","总金额(必填)","总实付(必填)","总优惠","邮费","标题(必填)","图片地址","货号","商家SKU","规格","数量(必填)","单价","优惠","实付(必填)","收件人(必填)","固话(必填)","手机(必填)","省份(必填)","城市(必填)","区县(必填)","详细地址(必填)","买家留言","卖家备注"]
    for k in data_dic:
        wb = Workbook()
        sheet = wb.active

        # 设置行高
        for i in range(3, len(data_dic[k])+1):
            sheet.row_dimensions[i].height = 61.5
        # 设置列宽
        column_dimensions = (20, 11.71, 12.14, 12.14, 5, 5, 38.14, 8.43, 5, 7.43, 4.86, 9.71, 3.86, 3.86, 9.14, 10.14, 11.14, 11.14, 9.14, 9.14, 9.14, 42.71,10, 10)
        for i in range(len(header)):
            sheet.column_dimensions[chr(65+i)].width = column_dimensions[i]


        sheet['A1']='''模版说明:
            \r\n1、标题行说明：【绿色为订单信息】，【蓝色为商品信息】；有标注【必填】的请按要求填写。
            \r\n2、一个订单多种商品，参考【灰色行】的示例数据（第3，4行是同一个订单；其中除了订单信息一样，商品信息不一样）
            \r\n3、一个订单一种商品，参考【粉红色行】的示例数据（第5行是一个订单）
            \r\n4、省市区请按标准国家地址库填写，否则会影响到电子面单中分拣信息的识别
            \r\n注意：请勿修改模版格式(第1,2行)，否则会导致导入失败！请严格按照要求认真填写数据后，再导入到软件中。
        '''
        sheet.merge_cells('A1:X1') # 合并一行中的几个单元格
        sheet.append(header)
        for data in data_dic[k]:
            sheet.append(data)
        file_name = r'%s_打印表.xlsx' % k
        out_path.append(file_name)
        wb.save( os.path.join(base_dir,file_name))
        wb.close()
    return out_path


def writex1_total_table(data_total_table, base_dir):
    header = ["订单号(必填)", "买家昵称(必填)", "总金额(必填)", "总实付(必填)", "总优惠", "邮费", "标题(必填)", "图片地址", "货号", "商家SKU", "规格", "数量(必填)",
              "单价", "优惠", "实付(必填)", "收件人(必填)", "固话(必填)", "手机(必填)", "省份(必填)", "城市(必填)", "区县(必填)", "详细地址(必填)", "买家留言",
              "卖家备注"]

    wb = Workbook()
    sheet = wb.active

    # 设置行高
    for i in range(3, len(data_total_table) + 1):
        sheet.row_dimensions[i].height = 61.5
    # 设置列宽
    column_dimensions = (
    20, 11.71, 12.14, 12.14, 5, 5, 38.14, 8.43, 5, 7.43, 4.86, 9.71, 3.86, 3.86, 9.14, 10.14, 11.14, 11.14, 9.14,
    9.14, 9.14, 42.71, 10, 10)
    for i in range(len(header)):
        sheet.column_dimensions[chr(65 + i)].width = column_dimensions[i]

    sheet['A1'] = '''模版说明:
            \r\n1、标题行说明：【绿色为订单信息】，【蓝色为商品信息】；有标注【必填】的请按要求填写。
            \r\n2、一个订单多种商品，参考【灰色行】的示例数据（第3，4行是同一个订单；其中除了订单信息一样，商品信息不一样）
            \r\n3、一个订单一种商品，参考【粉红色行】的示例数据（第5行是一个订单）
            \r\n4、省市区请按标准国家地址库填写，否则会影响到电子面单中分拣信息的识别
            \r\n注意：请勿修改模版格式(第1,2行)，否则会导致导入失败！请严格按照要求认真填写数据后，再导入到软件中。
        '''
    sheet.merge_cells('A1:X1')  # 合并一行中的几个单元格
    sheet.append(header)
    for data in data_total_table:
        sheet.append(data)
    file_name = r'%s_打印表.xlsx' % '总共'
    wb.save(os.path.join(base_dir, file_name))
    wb.close()
    return file_name

def main(in_file_path, base_dir):
    global SKU_STR, EXPRESS_REGION, EXPRESS_COUNT
    try:
        SKU_STR = load_config()
        # EXPRESS_REGION = load_express_region()
        EXPRESS_COUNT =load_express_count()
        # 按照地区的，暂时废弃
        deal_data, data_total_table = readxl_split_by_express_region(in_file_path)
        #按照数量的
        deal_data, data_total_table = readxl_split_by_express_by_count(in_file_path)
        out_file_paths = writexl(deal_data, base_dir)
        out_file_paths.append(writex1_total_table(data_total_table, base_dir))
        print('完成!')
        return out_file_paths
    except:
        print('Error!!!')
        return ['Error!!!']


# if __name__ == '__main__':
# # try:
#     print('完成!')
#     SKU_STR = load_config()
#     EXPRESS = load_express()
#     # deal_data = readxl(r'C:\Users\www\Desktop\bxs\1.xlsx')
#     deal_data = readxl(r'1.xlsx')
#     # print(deal_data)
#     writexl(deal_data)
#     print('完成!')
# # except:
# #     print('Error!!!')



